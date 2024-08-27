# one excel for all groups and points, split in files per group
import os
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup

# Load the Excel file with data_only=True to get the calculated values
excel_file = '../../__ignore/2024/Auswertung_2024.xlsx'
output_dir = '../../__ignore/2024/excel'
workbook = load_workbook(excel_file, data_only=True)
sheet = workbook.active

# Create a DataFrame from the loaded sheet
data = pd.DataFrame(sheet.values)

# Handle NaN values by replacing them with empty strings
data = data.fillna("")

# Function to get formatted value from cell
def get_formatted_value(cell):
    if cell is None or cell.value is None:
        return ""
    formatted_value = ""
    if cell.font.bold:
        formatted_value += "<b>"
    if cell.font.italic:
        formatted_value += "<i>"
    if cell.font.underline:
        formatted_value += "<u>"
    formatted_value += str(cell.value)
    if cell.font.underline:
        formatted_value += "</u>"
    if cell.font.italic:
        formatted_value += "</i>"
    if cell.font.bold:
        formatted_value += "</b>"
    return formatted_value

# Function to generate HTML document for each column E-T
def generate_html_docs(df):
    for col in range(4, 20):  # Columns E-T (0-indexed so 4-20)
        filename = os.path.join(output_dir, df.iloc[0, col] + '.html')
        with open(filename, 'w', encoding='utf-8') as file:
            html_content = '<table border="1">'
            html_content += "<tr><th></th><th>P. max</th><th>{}</th></tr>".format(df.iloc[0, col])
            for row in range(1, len(df)):
                cell_a = sheet.cell(row=row+1, column=1)
                cell_c = sheet.cell(row=row+1, column=3)
                cell_points = sheet.cell(row=row+1, column=col+1)
                html_content += "<tr>"
                html_content += "<td>{}</td>".format(get_formatted_value(cell_a))
                html_content += "<td>{}</td>".format(get_formatted_value(cell_c))
                html_content += "<td>{}</td>".format(get_formatted_value(cell_points))
                html_content += "</tr>"
            html_content += '</table>'
            soup = BeautifulSoup(html_content, 'html.parser')
            file.write(soup.prettify())

generate_html_docs(data)


# move all html files to __ignore/2024/excel folder